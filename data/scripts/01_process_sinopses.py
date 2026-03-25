"""
Processamento das Sinopses Estatísticas da Educação Básica (INEP).

Lê os arquivos Excel das Sinopses e extrai dados de Educação Profissional,
gerando JSONs processados em data/processed/ para consumo pelo site.

Uso:
    python 01_process_sinopses.py                    # processa todos os anos encontrados
    python 01_process_sinopses.py --anos 2023        # processa ano específico
    python 01_process_sinopses.py --verbose           # mostra detalhes do processamento

Estrutura da aba 1.30 das Sinopses (Matrículas EP por modalidade e rede):
- Linhas 1-10: cabeçalhos mesclados (título, subtítulos, nomes de colunas)
- Linha 11: dados Brasil (totais)
- Linhas 12+: dados por região, UF e município
- Colunas (45 total):
    0: Região, 1: UF, 2: Município, 3: Código Município, 4: Total geral
    5-9:   Técnico Integrado (Total, Federal, Estadual, Municipal, Privada)
    10-14: Normal/Magistério (Total, Federal, Estadual, Municipal, Privada)
    15-19: Concomitante (Total, Federal, Estadual, Municipal, Privada)
    20-24: Subsequente (Total, Federal, Estadual, Municipal, Privada)
    25-29: Integrada EJA (Total, Federal, Estadual, Municipal, Privada)
    30-34: FIC Concomitante (Total, Federal, Estadual, Municipal, Privada)
    35-39: FIC EJA Fundamental (Total, Federal, Estadual, Municipal, Privada)
    40-44: FIC EJA Médio (Total, Federal, Estadual, Municipal, Privada)
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

from reference_data import UFS, REGIOES

# Diretórios
RAW_DIR = Path(__file__).parent.parent / "raw"
PROCESSED_DIR = Path(__file__).parent.parent / "processed"

# Nomes das abas de Educação Profissional nas Sinopses
# A aba principal de matrículas é a que tem "1.30" ou "1.29" no nome
# (varia entre anos, mas sempre é a primeira aba de "Educação Profissional" na seção Matrículas)
EPT_SHEET_PATTERN = "Educação Profissional 1."

# Mapeamento de colunas da aba 1.30 (índices 0-based)
COL_REGIAO = 0
COL_UF = 1
COL_MUNICIPIO = 2
COL_COD_MUNICIPIO = 3
COL_TOTAL = 4

# Blocos de modalidade: (nome, índice_total, índice_federal, índice_estadual, índice_municipal, índice_privada)
MODALIDADES_COLUNAS = {
    "integrada": {"total": 5, "federal": 6, "estadual": 7, "municipal": 8, "privada": 9},
    "normal_magisterio": {"total": 10, "federal": 11, "estadual": 12, "municipal": 13, "privada": 14},
    "concomitante": {"total": 15, "federal": 16, "estadual": 17, "municipal": 18, "privada": 19},
    "subsequente": {"total": 20, "federal": 21, "estadual": 22, "municipal": 23, "privada": 24},
    "integrada_eja": {"total": 25, "federal": 26, "estadual": 27, "municipal": 28, "privada": 29},
    "fic_concomitante": {"total": 30, "federal": 31, "estadual": 32, "municipal": 33, "privada": 34},
    "fic_eja_fundamental": {"total": 35, "federal": 36, "estadual": 37, "municipal": 38, "privada": 39},
    "fic_eja_medio": {"total": 40, "federal": 41, "estadual": 42, "municipal": 43, "privada": 44},
}

# Modalidades "principais" de curso técnico (excluindo FIC e Normal/Magistério)
MODALIDADES_TECNICAS = ["integrada", "concomitante", "subsequente", "integrada_eja"]


def safe_int(value) -> int:
    """Converte valor para inteiro, tratando None, strings e formatos brasileiros."""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    if s in ("", "-", "--", "...", "—"):
        return 0
    # Remover separador de milhar (ponto) se houver
    s = s.replace(".", "").replace(",", "")
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def find_sinopse_files(raw_dir: Path) -> dict[int, Path]:
    """Encontra arquivos de Sinopse por ano em data/raw/ (incluindo subdiretórios)."""
    files = {}
    for xlsx in raw_dir.rglob("*.xlsx"):
        name = xlsx.name.lower()
        if "sinop" in name and "educa" in name:
            # Extrair ano do nome do arquivo
            for year in range(2015, 2030):
                if str(year) in name:
                    files[year] = xlsx
                    break
    return dict(sorted(files.items()))


def find_ept_matriculas_sheet(wb: openpyxl.Workbook) -> str | None:
    """Encontra a aba de matrículas de Educação Profissional."""
    for name in wb.sheetnames:
        if "Educação Profissional" in name and "1." in name:
            return name
    # Fallback: procurar por variações
    for name in wb.sheetnames:
        if "Profissional" in name and any(
            x in name for x in ["1.30", "1.29", "1.28", "1.31"]
        ):
            return name
    return None


def is_uf_row(row: tuple) -> bool:
    """Verifica se uma linha é de UF (tem região e UF mas não município)."""
    regiao = str(row[COL_REGIAO] or "").strip()
    uf = str(row[COL_UF] or "").strip()
    municipio = str(row[COL_MUNICIPIO] or "").strip()
    return bool(regiao) and bool(uf) and not municipio


def is_regiao_row(row: tuple) -> bool:
    """Verifica se uma linha é de região (tem região mas não UF nem município)."""
    regiao = str(row[COL_REGIAO] or "").strip()
    uf = str(row[COL_UF] or "").strip()
    municipio = str(row[COL_MUNICIPIO] or "").strip()
    return bool(regiao) and not uf and not municipio


def is_brasil_row(row: tuple) -> bool:
    """Verifica se uma linha é a linha Brasil."""
    regiao = str(row[COL_REGIAO] or "").strip().lower()
    return regiao.startswith("brasil")


def normalize_uf_name(name: str) -> str | None:
    """Normaliza nome de UF e retorna a sigla."""
    name = name.strip()
    for sigla, info in UFS.items():
        if info["nome"].lower() == name.lower():
            return sigla
    return None


def normalize_regiao_name(name: str) -> str | None:
    """Normaliza nome de região."""
    name = name.strip()
    for regiao in REGIOES:
        if regiao.lower() == name.lower():
            return regiao
        # Tratar "Centro-Oeste" vs "Centro Oeste"
        if regiao.lower().replace("-", " ") == name.lower().replace("-", " "):
            return regiao
    return None


def process_year(wb: openpyxl.Workbook, ano: int, verbose: bool = False) -> dict:
    """Processa um ano da Sinopse e retorna dicionário com os dados extraídos."""
    sheet_name = find_ept_matriculas_sheet(wb)
    if not sheet_name:
        print(f"  AVISO: Aba de Educação Profissional não encontrada para {ano}")
        return {}

    if verbose:
        print(f"  Aba encontrada: {sheet_name}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Dados de saída
    brasil_data = None
    regioes_data = []
    ufs_data = []

    # Processar linhas de dados (a partir da linha 11, índice 10)
    data_start = 10
    for i in range(data_start, len(rows)):
        row = rows[i]
        if len(row) < 25:
            continue

        total = safe_int(row[COL_TOTAL])
        if total == 0 and not is_brasil_row(row):
            # Pular municípios sem dados para agilizar (linhas de município com total=0)
            municipio = str(row[COL_MUNICIPIO] or "").strip()
            if municipio:
                continue

        if is_brasil_row(row):
            # Calcular totais por rede (somando todas as modalidades)
            federal = sum(safe_int(row[m["federal"]]) for m in MODALIDADES_COLUNAS.values() if m["federal"] < len(row))
            estadual = sum(safe_int(row[m["estadual"]]) for m in MODALIDADES_COLUNAS.values() if m["estadual"] < len(row))
            municipal = sum(safe_int(row[m["municipal"]]) for m in MODALIDADES_COLUNAS.values() if m["municipal"] < len(row))
            privada = sum(safe_int(row[m["privada"]]) for m in MODALIDADES_COLUNAS.values() if m["privada"] < len(row))

            brasil_data = {
                "ano": ano,
                "total": total,
                "federal": federal,
                "estadual": estadual,
                "municipal": municipal,
                "privada": privada,
                "publica": federal + estadual + municipal,
                "modalidades": {},
            }
            for mod_name, cols in MODALIDADES_COLUNAS.items():
                if cols["total"] < len(row):
                    brasil_data["modalidades"][mod_name] = safe_int(row[cols["total"]])

            if verbose:
                print(f"  Brasil: total={total:,}")

        elif is_regiao_row(row):
            regiao_name = normalize_regiao_name(str(row[COL_REGIAO] or ""))
            if regiao_name:
                regioes_data.append({
                    "ano": ano,
                    "regiao": regiao_name,
                    "total": total,
                    "modalidades": {
                        mod_name: safe_int(row[cols["total"]])
                        for mod_name, cols in MODALIDADES_COLUNAS.items()
                        if cols["total"] < len(row)
                    },
                })

        elif is_uf_row(row):
            uf_name = str(row[COL_UF] or "").strip()
            sigla = normalize_uf_name(uf_name)
            if sigla:
                federal = sum(safe_int(row[m["federal"]]) for m in MODALIDADES_COLUNAS.values() if m["federal"] < len(row))
                estadual = sum(safe_int(row[m["estadual"]]) for m in MODALIDADES_COLUNAS.values() if m["estadual"] < len(row))
                municipal = sum(safe_int(row[m["municipal"]]) for m in MODALIDADES_COLUNAS.values() if m["municipal"] < len(row))
                privada = sum(safe_int(row[m["privada"]]) for m in MODALIDADES_COLUNAS.values() if m["privada"] < len(row))

                ufs_data.append({
                    "ano": ano,
                    "sigla_uf": sigla,
                    "nome_uf": UFS[sigla]["nome"],
                    "regiao": UFS[sigla]["regiao"],
                    "total": total,
                    "federal": federal,
                    "estadual": estadual,
                    "municipal": municipal,
                    "privada": privada,
                    "publica": federal + estadual + municipal,
                    "modalidades": {
                        mod_name: safe_int(row[cols["total"]])
                        for mod_name, cols in MODALIDADES_COLUNAS.items()
                        if cols["total"] < len(row)
                    },
                })

    if verbose:
        print(f"  Regiões: {len(regioes_data)}, UFs: {len(ufs_data)}")

    return {
        "brasil": brasil_data,
        "regioes": regioes_data,
        "ufs": ufs_data,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Processamento das Sinopses Estatísticas do INEP"
    )
    parser.add_argument(
        "--anos", nargs="+", type=int, default=None,
        help="Anos para processar (default: todos encontrados)",
    )
    parser.add_argument("--verbose", action="store_true", help="Mostrar detalhes")
    args = parser.parse_args()

    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("Processamento - Sinopses Estatísticas EPT (INEP)")
    print("=" * 60)

    # Encontrar arquivos
    sinopse_files = find_sinopse_files(RAW_DIR)
    if not sinopse_files:
        print(f"\nNenhum arquivo de Sinopse encontrado em {RAW_DIR.resolve()}")
        print("Execute primeiro: python 00_download_sinopses.py")
        return 1

    print(f"\nArquivos encontrados: {list(sinopse_files.keys())}")

    if args.anos:
        sinopse_files = {k: v for k, v in sinopse_files.items() if k in args.anos}

    # Processar cada ano
    all_brasil = []
    all_regioes = []
    all_ufs = []

    for ano, filepath in sorted(sinopse_files.items()):
        print(f"\n--- Processando {ano}: {filepath.name} ---")
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            result = process_year(wb, ano, verbose=args.verbose)
            wb.close()

            if result.get("brasil"):
                all_brasil.append(result["brasil"])
                all_regioes.extend(result["regioes"])
                all_ufs.extend(result["ufs"])
                print(f"  OK: {result['brasil']['total']:,} matrículas totais")
            else:
                print(f"  AVISO: Sem dados extraídos para {ano}")

        except Exception as e:
            print(f"  ERRO ao processar {ano}: {e}")
            if args.verbose:
                import traceback
                traceback.print_exc()

    if not all_brasil:
        print("\nNenhum dado processado. Verifique os arquivos.")
        return 1

    # Gerar JSONs de saída
    print(f"\n{'=' * 60}")
    print("Gerando JSONs...")

    # 1. Série temporal (totais por ano)
    serie_temporal = []
    for b in sorted(all_brasil, key=lambda x: x["ano"]):
        serie_temporal.append({
            "ano": b["ano"],
            "total": b["total"],
            "publica": b["publica"],
            "privada": b["privada"],
            "federal": b["federal"],
            "estadual": b["estadual"],
            "municipal": b["municipal"],
        })

    write_json(PROCESSED_DIR / "matriculas_ept_serie_temporal.json", serie_temporal)

    # 2. Matrículas por UF
    write_json(PROCESSED_DIR / "matriculas_ept_por_uf.json", all_ufs)

    # 3. Matrículas por modalidade
    modalidades_serie = []
    for b in sorted(all_brasil, key=lambda x: x["ano"]):
        for mod_name in MODALIDADES_TECNICAS:
            modalidades_serie.append({
                "ano": b["ano"],
                "modalidade": mod_name,
                "total": b["modalidades"].get(mod_name, 0),
            })

    write_json(PROCESSED_DIR / "matriculas_ept_por_modalidade.json", modalidades_serie)

    # 4. Matrículas por região
    write_json(PROCESSED_DIR / "matriculas_ept_por_regiao.json", all_regioes)

    # 5. Metadata
    metadata = {
        "fonte": "Sinopses Estatísticas da Educação Básica - INEP/MEC",
        "url_fonte": "https://www.gov.br/inep/pt-br/acesso-a-informacao/dados-abertos/sinopses-estatisticas",
        "data_processamento": datetime.now().isoformat(timespec="seconds"),
        "anos_cobertos": sorted([b["ano"] for b in all_brasil]),
        "total_registros_uf": len(all_ufs),
        "total_registros_regiao": len(all_regioes),
        "modalidades_incluidas": MODALIDADES_TECNICAS,
        "notas": [
            "Dados extraídos da aba de Educação Profissional das Sinopses Estatísticas",
            "Modalidades técnicas: integrada, concomitante, subsequente, integrada_eja",
            "Redes: federal, estadual, municipal, privada",
            "Pública = federal + estadual + municipal",
        ],
    }

    write_json(PROCESSED_DIR / "metadata.json", metadata)

    # Resumo
    print(f"\n{'=' * 60}")
    print("Resumo:")
    print(f"  Anos processados: {metadata['anos_cobertos']}")
    print(f"  Registros UF: {metadata['total_registros_uf']}")
    print(f"  Registros região: {metadata['total_registros_regiao']}")
    for b in sorted(all_brasil, key=lambda x: x["ano"]):
        print(f"  {b['ano']}: {b['total']:>12,} matrículas ({b['publica']:,} pública, {b['privada']:,} privada)")
    print(f"\nArquivos gerados em: {PROCESSED_DIR.resolve()}")
    print("=" * 60)

    return 0


def write_json(path: Path, data):
    """Escreve dados como JSON formatado."""
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    size_kb = path.stat().st_size / 1024
    print(f"  {path.name} ({size_kb:.1f} KB)")


if __name__ == "__main__":
    sys.exit(main())
